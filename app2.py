from flask import Flask, render_template, request
import openpyxl
import pandas as pd
import os

app = Flask(__name__)
workbook = openpyxl.load_workbook('./Nikao_Data/score_search/Scoresheet_2022.xlsx')

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/score', methods=['POST'])
def score():
    # Get the sheet names representing each month
    months = workbook.sheetnames

    # Get user input for their name, day, and month
    name = request.form['name']
    day = request.form['day']
    month = request.form['month']

    # Check if the entered month is valid
    if month not in months:
        return "Oops! That didn't work. Try again"
    else:
        # get the worksheet for the entered month
        sheet = workbook[month]

        # Loop through the cells in the first row to find the column with the entered day
        for cell in sheet[1]:
            if cell.value == day:
                col_index = cell.column_letter
                break

        # Loop through the cells in column A to find the row with the entered name
        for row in sheet.iter_rows(min_row=2, min_col=1):
            if row[0].value == name:
                row_index = row[0].row
                break

        # Get the value of the cell at the intersection of the row and column
        value = sheet[col_index + str(row_index)].value

        # Render the template with the dynamic content
        return render_template('score.html', name=name, month=month, day=day, score=value)

if __name__ == '__main__':
    app.run(debug=True)
